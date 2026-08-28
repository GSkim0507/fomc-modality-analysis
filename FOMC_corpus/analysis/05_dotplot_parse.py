"""
Parse Bloomberg dot-plot xlsx files into a tidy CSV of median projections.

Bloomberg files come in two shapes:
  A) Single-meeting dot plot:
        sheet[0] title: "FOMC 점도표 MM-DD-YYYY"  (distribution)
        sheet[1] title: "곡선"                     (median + OIS, columns = years/'장기')
  B) Multi-meeting history:
        sheet[0] title: "DOTS 변화 YYYY"           (distribution across meeting dates)
        sheet[1] title: "곡선"                     (medians for each meeting date)

We pull "FOMC 점도표 중간값" (median) rows. The year-target columns differ between
A and B (A: projection horizon year; B: meeting date). We unify into rows:
  meeting_date (ISO), projection_horizon (year or 'longer_run'), median_pct

Output: analysis/dotplot_medians.csv
        analysis/dotplot_labels.csv  -- per meeting: hawkish/dovish/neutral label
"""
from __future__ import annotations
import csv
import re
from collections import defaultdict
from pathlib import Path
from datetime import datetime

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
BLOOMBERG_DIR = Path("/Users/gohyeonjeong/Documents/hz/FOMC 블룸버그")
ANALYSIS = ROOT / "analysis"

# Sheet title patterns
RE_SINGLE_TITLE = re.compile(r"(\d{1,2})[-./](\d{1,2})[-./](\d{4})")    # e.g. 12-14-2022
RE_HIST_TITLE = re.compile(r"DOTS\s*변화\s*(\d{4})", re.I)
LONG_RUN_LABELS = {"장기", "longer run", "long run", "lr"}
MEDIAN_ROW_LABEL = "FOMC 점도표 중간값"

def parse_curve_columns(headers: list) -> list[tuple[str, str]]:
    """
    Given a header row from the '곡선' sheet, return list of (column_kind, value) tuples
    where column_kind is one of:
      - 'year'         (value = '2022', '2023', ...)
      - 'longer_run'
      - 'meeting_date' (value = ISO date string)
      - 'desc'         (descriptive column, skip)
    """
    out = []
    for h in headers:
        s = "" if h is None else str(h).strip()
        if not s:
            out.append(("skip", ""))
            continue
        sl = s.lower()
        # Year column like '2022'
        if re.fullmatch(r"20\d{2}", s):
            out.append(("year", s))
            continue
        # Longer run
        if any(k in s for k in LONG_RUN_LABELS):
            out.append(("longer_run", ""))
            continue
        # Meeting date like 12/14/22 or 12/14/2022
        m = re.fullmatch(r"(\d{1,2})/(\d{1,2})/(\d{2,4})", s)
        if m:
            mo, da, yr = int(m.group(1)), int(m.group(2)), int(m.group(3))
            if yr < 100: yr += 2000
            iso = f"{yr:04d}-{mo:02d}-{da:02d}"
            out.append(("meeting_date", iso))
            continue
        # Descriptive (e.g. '곡선', 'ti/p/ol')
        out.append(("desc", s))
    return out

def coerce_pct(v) -> float | None:
    if v is None or v == "": return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None

def parse_workbook(p: Path) -> list[dict]:
    """Return list of {meeting_date, projection_horizon, median_pct, source_file}."""
    out = []
    try:
        wb = load_workbook(p, read_only=True, data_only=True)
    except Exception as e:
        print(f"  ERROR opening {p.name}: {e}")
        return out

    sheets = wb.worksheets
    if not sheets:
        return out

    # Figure out the "meeting date" associated with this file
    file_meeting_date = None
    file_history_year = None
    for ws in sheets:
        t = ws.title
        m_single = RE_SINGLE_TITLE.search(t)
        if m_single and "점도표" in t:
            mo, da, yr = int(m_single.group(1)), int(m_single.group(2)), int(m_single.group(3))
            file_meeting_date = f"{yr:04d}-{mo:02d}-{da:02d}"
        m_hist = RE_HIST_TITLE.search(t)
        if m_hist:
            file_history_year = m_hist.group(1)

    # Pull the 곡선 sheet (typically last)
    curve = next((ws for ws in sheets if "곡선" in ws.title), None)
    if curve is None:
        return out

    rows = list(curve.iter_rows(values_only=True))
    if not rows:
        return out

    # Header detection: find row with at least one year column OR meeting_date column
    header_idx = None
    parsed_headers = None
    for i, row in enumerate(rows[:5]):
        ph = parse_curve_columns(list(row))
        kinds = {k for k, _ in ph}
        if kinds & {"year", "meeting_date", "longer_run"}:
            header_idx = i
            parsed_headers = ph
            break
    if parsed_headers is None:
        return out

    # Find the median row
    for row in rows[header_idx + 1:]:
        first = (row[0] if row else None)
        if first is None: continue
        if MEDIAN_ROW_LABEL not in str(first): continue
        # Pull values
        for col_idx, (kind, val) in enumerate(parsed_headers):
            if col_idx >= len(row): break
            cell = row[col_idx]
            if kind == "year":
                pct = coerce_pct(cell)
                if pct is None: continue
                out.append({
                    "meeting_date": file_meeting_date or "",
                    "projection_horizon": val,        # '2022', '2023', ...
                    "median_pct": pct,
                    "source_file": p.name,
                    "source_kind": "single" if file_meeting_date else "history",
                    "history_year": file_history_year or "",
                })
            elif kind == "longer_run":
                pct = coerce_pct(cell)
                if pct is None: continue
                out.append({
                    "meeting_date": file_meeting_date or "",
                    "projection_horizon": "longer_run",
                    "median_pct": pct,
                    "source_file": p.name,
                    "source_kind": "single" if file_meeting_date else "history",
                    "history_year": file_history_year or "",
                })
            elif kind == "meeting_date":
                pct = coerce_pct(cell)
                if pct is None: continue
                out.append({
                    "meeting_date": val,                  # column header IS the meeting date
                    "projection_horizon": file_history_year or "",
                    "median_pct": pct,
                    "source_file": p.name,
                    "source_kind": "history",
                    "history_year": file_history_year or "",
                })
        break  # only first median row
    return out

def main():
    files = sorted(BLOOMBERG_DIR.glob("grid1_*.xlsx"))
    print(f"Found {len(files)} dotplot files")

    rows = []
    for p in files:
        recs = parse_workbook(p)
        print(f"  {p.name}: {len(recs)} median rows")
        rows.extend(recs)

    # Dedupe: prefer 'single' source (more authoritative) over 'history' if same (meeting_date, horizon)
    dedup: dict[tuple, dict] = {}
    for r in rows:
        key = (r["meeting_date"], r["projection_horizon"])
        if key not in dedup:
            dedup[key] = r
        else:
            # prefer single
            if r["source_kind"] == "single" and dedup[key]["source_kind"] != "single":
                dedup[key] = r
    rows = sorted(dedup.values(), key=lambda r: (r["meeting_date"], str(r["projection_horizon"])))

    out_csv = ANALYSIS / "dotplot_medians.csv"
    cols = ["meeting_date", "projection_horizon", "median_pct", "source_file", "source_kind", "history_year"]
    with out_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader(); w.writerows(rows)
    print(f"\nwrote {out_csv}: {len(rows)} rows")

    # Build labels: for each meeting, compare current-year (or nearest projection-year) median
    # to PRIOR meeting's median for the SAME projection year.
    # Strategy: for each meeting m_i, pick projection_horizon = year(m_i)+1 if exists, else year(m_i).
    # Compare to m_{i-1}'s median for the same horizon. delta>0 -> hawkish; <0 -> dovish; ==0 -> neutral.
    meetings = sorted({r["meeting_date"] for r in rows if r["meeting_date"]})
    by_meeting_horizon: dict[tuple, float] = {(r["meeting_date"], r["projection_horizon"]): r["median_pct"] for r in rows}

    labels = []
    for i, m in enumerate(meetings):
        if i == 0:
            labels.append({"meeting_date": m, "label": "baseline",
                           "delta_pct": None, "horizon": "", "prior_meeting": ""})
            continue
        prev = meetings[i - 1]
        m_yr = int(m[:4])
        # Try horizons in order: current+1, current, current-1
        chosen = None
        for h in (str(m_yr + 1), str(m_yr), str(m_yr - 1), "longer_run"):
            if (m, h) in by_meeting_horizon and (prev, h) in by_meeting_horizon:
                chosen = h; break
        if chosen is None:
            labels.append({"meeting_date": m, "label": "no_comparable",
                           "delta_pct": None, "horizon": "", "prior_meeting": prev})
            continue
        delta = by_meeting_horizon[(m, chosen)] - by_meeting_horizon[(prev, chosen)]
        # Label cutoffs: ≥+0.125pp = hawkish (one dot-step), ≤-0.125pp = dovish
        if delta >= 0.125: lab = "hawkish"
        elif delta <= -0.125: lab = "dovish"
        else: lab = "neutral"
        labels.append({"meeting_date": m, "label": lab,
                       "delta_pct": round(delta, 4),
                       "horizon": chosen, "prior_meeting": prev})

    out_lbl = ANALYSIS / "dotplot_labels.csv"
    with out_lbl.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["meeting_date","label","delta_pct","horizon","prior_meeting"])
        w.writeheader(); w.writerows(labels)
    print(f"wrote {out_lbl}: {len(labels)} labeled meetings")
    from collections import Counter
    print("label distribution:", Counter(l["label"] for l in labels))
    print(f"meetings with comparable delta: {sum(1 for l in labels if l['label'] in {'hawkish','dovish','neutral'})}")

if __name__ == "__main__":
    main()
